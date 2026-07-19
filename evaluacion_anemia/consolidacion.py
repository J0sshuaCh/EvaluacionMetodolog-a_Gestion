"""
consolidacion.py - Exportación final a Excel
==============================================
Genera el archivo evaluacion_resultados_final.xlsx con 3 hojas:

1. DETALLE: Cada pregunta fila con todas las métricas
2. RESUMEN_CATEGORIAS: Promedios por categoría (Prevención, Diagnóstico, Tratamiento)
3. COMPARATIVA: Sin RAG vs Con RAG en cada métrica
"""

import os
import numpy as np
import pandas as pd
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

from config import OUTPUT_DIR


# ── Estilos ──────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="2F5496")
SCORE_FONT = Font(name="Calibri", size=11)
MEJOR_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PEOR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Columnas a incluir en cada hoja ─────────────────────────────

COLUMNAS_DETALLE = [
    # Identificación
    "ID", "Categoria", "Pregunta",
    # Respuestas
    "Respuesta_Referencia_Ground_Truth",
    "Respuesta_Sin_RAG",
    "Respuesta_Con_RAG",
    "Contexto_Recuperado",
    # Groundedness
    "Groundedness_ConRAG",
    "Faithfulness_Afirmaciones",
    "Faithfulness_Respaldadas",
    # Concordancia
    "Concordancia_SinRAG",
    "Concordancia_ConRAG",
    "Concordancia_Semantica_SinRAG",
    "Concordancia_Semantica_ConRAG",
    "Concordancia_Entidades_SinRAG",
    "Concordancia_Entidades_ConRAG",
    "Precision_Factual_SinRAG",
    "Precision_Factual_ConRAG",
    "Seguridad_SinRAG",
    "Seguridad_ConRAG",
    # Opacidad Epistémica
    "Opacidad_Score_SinRAG",
    "Opacidad_Score_ConRAG",
    "Opacidad_Clasificacion_SinRAG",
    "Opacidad_Clasificacion_ConRAG",
    "Opacidad_Densidad_SinRAG",
    "Opacidad_Densidad_ConRAG",
    # Potencial de Engaño
    "PoD_Score_SinRAG",
    "PoD_Score_ConRAG",
    "PoD_Ajustado_SinRAG",
    "PoD_Ajustado_ConRAG",
    "PoD_Clasificacion_SinRAG",
    "PoD_Clasificacion_ConRAG",
    "PoD_Persuasividad_SinRAG",
    "PoD_Persuasividad_ConRAG",
]


def _aplicar_estilos(ws, df: pd.DataFrame):
    """Aplica formato a las celdas de una hoja."""
    # Encabezados
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Filas de datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = SCORE_FONT
            cell.alignment = LEFT_ALIGN if cell.column <= 3 else CENTER_ALIGN
            cell.border = THIN_BORDER

    # Ajustar ancho de columnas
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[chr(64 + col_idx) if col_idx < 27 else "A"].width = (
            min(max(max_len + 2, 12), 50)
        )

    # Alternar colores de filas
    light_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                if cell.column <= 3:
                    cell.fill = light_fill

    # Resaltar scores críticos en columnas de métricas
    # (a partir de columna index)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _estilo_titulo(ws, row, col, texto, level="title"):
    """Escribe un título con estilo."""
    cell = ws.cell(row=row, column=col, value=texto)
    cell.font = TITLE_FONT if level == "title" else SUBTITLE_FONT
    cell.alignment = LEFT_ALIGN
    return row + 1


def _estilo_par(ws, row, col, label, value, formato=".3f"):
    """Escribe una etiqueta y su valor."""
    ws.cell(row=row, column=col, value=label).font = Font(name="Calibri", bold=True, size=11)
    if isinstance(value, float):
        ws.cell(row=row, column=col + 1, value=round(value, 4)).number_format = formato
    else:
        ws.cell(row=row, column=col + 1, value=str(value))
    return row + 1


def _resumen_estadisticas(datos: pd.Series) -> Dict:
    """Calcula estadísticas descriptivas de una serie."""
    return {
        "media": datos.mean(),
        "mediana": datos.median(),
        "min": datos.min(),
        "max": datos.max(),
        "std": datos.std(),
        "count": len(datos),
    }


# ====================================================================
# Hoja 1: Detalle
# ====================================================================

def crear_hoja_detalle(wb: Workbook, df: pd.DataFrame):
    """Crea la hoja de detalle con todas las filas y métricas."""
    ws = wb.active
    ws.title = "Detalle"

    # Seleccionar solo columnas que existen en el DataFrame
    cols_disponibles = [c for c in COLUMNAS_DETALLE if c in df.columns]

    # Si faltan columnas clave, usar las que hay
    df_detalle = df[cols_disponibles].copy()

    # Escribir DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df_detalle, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER

    _aplicar_estilos(ws, df_detalle)


# ====================================================================
# Hoja 2: Resumen por categorías
# ====================================================================

def crear_hoja_resumen(wb: Workbook, df: pd.DataFrame):
    """Crea hoja con promedios por categoría."""
    ws = wb.create_sheet("Resumen Categorias")

    if "Categoria" not in df.columns:
        ws.cell(row=1, column=1, value="Columna 'Categoria' no encontrada")
        return

    # Definir qué métricas promediar
    metricas_promedio = [
        "Groundedness_ConRAG",
        "Concordancia_SinRAG", "Concordancia_ConRAG",
        "Concordancia_Semantica_SinRAG", "Concordancia_Semantica_ConRAG",
        "Concordancia_Entidades_SinRAG", "Concordancia_Entidades_ConRAG",
        "Precision_Factual_SinRAG", "Precision_Factual_ConRAG",
        "Seguridad_SinRAG", "Seguridad_ConRAG",
        "Opacidad_Score_SinRAG", "Opacidad_Score_ConRAG",
        "PoD_Ajustado_SinRAG", "PoD_Ajustado_ConRAG",
        "PoD_Persuasividad_SinRAG", "PoD_Persuasividad_ConRAG",
    ]
    metricas_disponibles = [m for m in metricas_promedio if m in df.columns]

    if not metricas_disponibles:
        ws.cell(row=1, column=1, value="No hay métricas disponibles para resumir")
        return

    # Agrupar por categoría (excluir preguntas mal clasificadas)
    categorias_raw = df["Categoria"].dropna().unique()
    # Filtrar categorías que no parezcan preguntas (demasiado largas)
    categorias = sorted([
        c for c in categorias_raw
        if len(str(c)) < 80  # Las preguntas son largas, las categorías cortas
    ])
    if not categorias:
        # Fallback: todas
        categorias = sorted(categorias_raw)

    # ── Tabla de promedios ──
    fila = 1
    fila = _estilo_titulo(ws, fila, 1, "Promedio de Metricas por Categoria")

    # Encabezados
    fila += 1
    ws.cell(row=fila, column=1, value="Métrica").font = HEADER_FONT
    ws.cell(row=fila, column=1).fill = HEADER_FILL
    for j, cat in enumerate(categorias, 2):
        ws.cell(row=fila, column=j, value=cat).font = HEADER_FONT
        ws.cell(row=fila, column=j).fill = HEADER_FILL
    ws.cell(row=fila, column=len(categorias) + 2, value="Global").font = HEADER_FONT
    ws.cell(row=fila, column=len(categorias) + 2).fill = HEADER_FILL
    for c in range(1, len(categorias) + 3):
        ws.cell(row=fila, column=c).border = THIN_BORDER
        ws.cell(row=fila, column=c).alignment = CENTER_ALIGN

    fila += 1
    for metrica in metricas_disponibles:
        ws.cell(row=fila, column=1, value=metrica).font = Font(name="Calibri", bold=True)
        ws.cell(row=fila, column=1).border = THIN_BORDER
        ws.cell(row=fila, column=1).alignment = LEFT_ALIGN

        for j, cat in enumerate(categorias, 2):
            valores = df[df["Categoria"] == cat][metrica].dropna()
            valor = valores.mean() if len(valores) > 0 else 0
            cell = ws.cell(row=fila, column=j, value=round(valor, 4))
            cell.number_format = "0.0000"
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            # Colorear según valor
            if "PoD" in metrica or "Opacidad" in metrica:
                if valor > 0.5:
                    cell.fill = PEOR_FILL
                elif valor < 0.15:
                    cell.fill = MEJOR_FILL
            else:
                if valor < 0.3:
                    cell.fill = PEOR_FILL
                elif valor > 0.7:
                    cell.fill = MEJOR_FILL

        # Global
        val_global = df[metrica].mean()
        cell = ws.cell(row=fila, column=len(categorias) + 2, value=round(val_global, 4))
        cell.number_format = "0.0000"
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        cell.font = Font(name="Calibri", bold=True)

        fila += 1

    ws.column_dimensions["A"].width = 40
    for j in range(2, len(categorias) + 3):
        ws.column_dimensions[chr(64 + j) if j < 27 else "A"].width = 22

    ws.auto_filter.ref = f"A1:{chr(64 + len(categorias) + 2)}{fila}"
    ws.freeze_panes = "B2"


# ====================================================================
# Hoja 3: Comparativa Sin RAG vs Con RAG
# ====================================================================

def crear_hoja_comparativa(wb: Workbook, df: pd.DataFrame):
    """Crea hoja comparativa entre escenarios."""
    ws = wb.create_sheet("Comparativa RAG")

    fila = 1
    fila = _estilo_titulo(ws, fila, 1, "📈 Comparativa: Sin RAG vs Con RAG")

    # Métricas a comparar (pares: columna_sin, columna_con, nombre_mostrar)
    pares_metricas = [
        ("Concordancia_SinRAG", "Concordancia_ConRAG", "Concordancia con Directrices"),
        ("Concordancia_Semantica_SinRAG", "Concordancia_Semantica_ConRAG",
         "  └─ Componente Semántico"),
        ("Concordancia_Entidades_SinRAG", "Concordancia_Entidades_ConRAG",
         "  └─ Coincidencia de Entidades"),
        ("Precision_Factual_SinRAG", "Precision_Factual_ConRAG", "Precisión Factual"),
        ("Seguridad_SinRAG", "Seguridad_ConRAG", "Seguridad"),
        ("Opacidad_Score_SinRAG", "Opacidad_Score_ConRAG", "Opacidad Epistémica (↓ mejor)"),
        ("PoD_Ajustado_SinRAG", "PoD_Ajustado_ConRAG", "Potencial de Engaño (↓ mejor)"),
        ("PoD_Persuasividad_SinRAG", "PoD_Persuasividad_ConRAG",
         "  └─ Persuasividad (↓ mejor)"),
    ]

    # Encabezados
    fila += 1
    headers = ["Métrica", "Sin RAG", "Con RAG", "Diferencia", "Mejora %", "Ganador"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=fila, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    fila += 1
    metricas_validas = []
    for col_sin, col_con, nombre in pares_metricas:
        if col_sin in df.columns and col_con in df.columns:
            media_sin = df[col_sin].mean()
            media_con = df[col_con].mean()
            diff = media_con - media_sin

            # Mejora porcentual (interpretación según métrica)
            if "Opacidad" in col_sin or "PoD" in col_sin or "Persuasividad" in col_sin:
                # ↓ mejor: mejora si disminuye
                mejora_pct = ((media_sin - media_con) / max(abs(media_sin), 0.001)) * 100
                ganador = "✅ Con RAG" if media_con < media_sin else "❌ Sin RAG"
                if abs(media_con - media_sin) < 0.01:
                    ganador = "⚖️ Empate"
            else:
                # ↑ mejor: mejora si aumenta
                mejora_pct = ((media_con - media_sin) / max(abs(media_sin), 0.001)) * 100
                ganador = "✅ Con RAG" if media_con > media_sin else "❌ Sin RAG"
                if abs(media_con - media_sin) < 0.01:
                    ganador = "⚖️ Empate"

            metricas_validas.append({
                "nombre": nombre,
                "sin_rag": media_sin,
                "con_rag": media_con,
                "diff": diff,
                "mejora_pct": mejora_pct,
                "ganador": ganador,
            })

            ws.cell(row=fila, column=1, value=nombre).font = Font(name="Calibri", bold=True)
            ws.cell(row=fila, column=1).border = THIN_BORDER

            for j, val in enumerate([media_sin, media_con, diff], 2):
                cell = ws.cell(row=fila, column=j, value=round(val, 4))
                cell.number_format = "0.0000"
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN

            ws.cell(row=fila, column=5, value=f"{mejora_pct:+.1f}%").border = THIN_BORDER
            ws.cell(row=fila, column=5).alignment = CENTER_ALIGN

            cell = ws.cell(row=fila, column=6, value=ganador)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            if "Con RAG" in ganador:
                cell.fill = MEJOR_FILL
            elif "Sin RAG" in ganador:
                cell.fill = PEOR_FILL

            fila += 1

    # ── Resumen general ──
    fila += 2
    fila = _estilo_titulo(ws, fila, 1, "🏆 Balance General", level="sub")

    # Contar ganadores
    ganados_con = sum(1 for m in metricas_validas if "Con RAG" in m["ganador"])
    ganados_sin = sum(1 for m in metricas_validas if "Sin RAG" in m["ganador"])
    empates = sum(1 for m in metricas_validas if "Empate" in m["ganador"])

    fila = _estilo_par(ws, fila, 1, "Métricas donde gana Con RAG:", ganados_con)
    fila = _estilo_par(ws, fila, 1, "Métricas donde gana Sin RAG:", ganados_sin)
    fila = _estilo_par(ws, fila, 1, "Empates:", empates)

    fila += 1
    total_metricas = len(metricas_validas)
    pct_mejora = (ganados_con / total_metricas * 100) if total_metricas > 0 else 0
    ws.cell(
        row=fila, column=1,
        value=f"La arquitectura RAG mejora en {pct_mejora:.0f}% de las métricas evaluadas."
    ).font = Font(name="Calibri", bold=True, size=12, color="2F5496")

    # Anchos de columna
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    ws.freeze_panes = "A2"


# ====================================================================
# Pipeline principal
# ====================================================================

def exportar_excel(
    df: pd.DataFrame,
    nombre_archivo: str = "evaluacion_resultados_final.xlsx",
) -> str:
    """Genera el Excel final con las 3 hojas.

    Args:
        df: DataFrame completo con todas las métricas.
        nombre_archivo: Nombre del archivo de salida.

    Returns:
        Ruta completa al archivo generado.
    """
    print("=" * 60)
    print("📁 EXPORTACIÓN A EXCEL")
    print("=" * 60)

    ruta_salida = os.path.join(OUTPUT_DIR, nombre_archivo)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()

    # Hoja 1: Detalle
    print("   Creando hoja Detalle...")
    crear_hoja_detalle(wb, df)

    # Hoja 2: Resumen por categorías
    print("   Creando hoja Resumen por Categorías...")
    crear_hoja_resumen(wb, df)

    # Hoja 3: Comparativa
    print("   Creando hoja Comparativa RAG...")
    crear_hoja_comparativa(wb, df)

    # Guardar
    wb.save(ruta_salida)
    print(f"\n✅ Excel generado: {ruta_salida}")
    print(f"   Hojas:")
    print(f"   1. Detalle — {len(df)} filas con todas las métricas")
    print(f"   2. Resumen Categorías — Promedios por grupo temático")
    print(f"   3. Comparativa RAG — Sin RAG vs Con RAG")

    return ruta_salida


if __name__ == "__main__":
    # Prueba con datos ficticios
    print("🔬 Módulo de exportación listo.")
    print("   Usar desde main.py para generar el Excel final.")
